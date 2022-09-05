import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
        }

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))

print(results)

try:
    wb = load_workbook(r"13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save(r"13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx")
