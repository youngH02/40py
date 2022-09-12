#pip install openpyxl #엑셀을 사용하기 위함
#pip install python-docx #워드 사용
#pip install docx2pdf #워드를 pdf로 변환

from math import degrees
from openpyxl import load_workbook
import os

os.chdir('/Users/jyoung/study/40py/12.엑셀 정보를 불러와 수료증 자동생성/')
load_wb = load_workbook(r'studentList.xlsx')
load_wb = load_wb.active

for i in range(1, load_wb.max_row+1) :
  print(load_wb.cell(i,1).value)