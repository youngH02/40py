import webbrowser
import requests
import re
from openpyxl import load_workbook, Workbook

test_string = """
aaa@bbb.com
123@abc.co.kr
test@hello.kr
ok@ok.co.kr
ok@ok.co.kr
ok@ok.co.kr
no.co.kr
no.kr
"""

url = 'https://www.newdaily.co.kr/site/data/html/2022/09/05/2022090500148.html'
headers = {
  'User-Agent' : 'Mozilla/5.0',
  'Content-Type': 'text/html; charset=utf-8'
}
response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w-]+\.[\w\.-]+',response.text)
results = set(results)
print(results)

try :
  wb = load_workbook(r'13.이메일 수집하여 엑셀에 기록/email_list.xlsx', data_only=True)
  sheet = wb.active
except :
  wb=Workbook()
  sheet = wb.active

for result in results :
  sheet.append([result])

wb.save(r'13.이메일 수집하여 엑셀에 기록/email_list.xlsx')

for i in range(1,sheet.max_row + 1):
  print(sheet.cell(i, 1).value)